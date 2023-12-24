#Код читает файл xml по ячейкам, считает количиство слогов в тексте ячейки(RU+EN).
# -*- coding: cp1251 -*- 
import openpyxl
import re
from langdetect import detect
import pyphen

# Открываем файл Excel
workbook = openpyxl.load_workbook("test.xlsx") #XML file read
sheet = workbook.active

def sylco(word):
    word = word.lower()
    exception_add = ['serious', 'crucial']
    exception_del = ['fortunately', 'unfortunately']
    co_one = ['cool', 'coach', 'coat', 'coal', 'count', 'coin', 'coarse', 'coup', 'coif', 'cook', 'coign', 'coiffe', 'coof', 'court']
    co_two = ['coapt', 'coed', 'coinci']
    pre_one = ['preach']
    syls = 0
    disc = 0
    
    #Правила првоерки слогов в EN
    # 1)
    if len(word) <= 3:
        syls = 1
        return syls

    # 2)
    if word[-2:] == "es" or word[-2:] == "ed":
        doubleAndtripple_1 = len(re.findall(r'[eaoui][eaoui]', word))
        if doubleAndtripple_1 > 1 or len(re.findall(r'[eaoui][^eaoui]', word)) > 1:
            if word[-3:] == "ted" or word[-3:] == "tes" or word[-3:] == "ses" or word[-3:] == "ied" or word[-3:] == "ies":
                pass
            else:
                disc += 1

    # 3)
    le_except = ['whole', 'mobile', 'pole', 'male', 'female', 'hale', 'pale', 'tale', 'sale', 'aisle', 'whale', 'while']
    if word[-1:] == "e":
        if word[-2:] == "le" and word not in le_except:
            pass
        else:
            disc += 1

    # 4)
    doubleAndtripple = len(re.findall(r'[eaoui][eaoui]', word))
    tripple = len(re.findall(r'[eaoui][eaoui][eaoui]', word))
    disc += doubleAndtripple + tripple

    # 5)
    numVowels = len(re.findall(r'[eaoui]', word))

    # 6)
    if word[:2] == "mc":
        syls += 1

    # 7)
    if word[-1:] == "y" and word[-2] not in "aeoui":
        syls += 1

    # 8)
    for i, j in enumerate(word):
        if j == "y":
            if (i != 0) and (i != len(word)-1):
                if word[i-1] not in "aeoui" and word[i+1] not in "aeoui":
                    syls += 1

    # 9)
    if word[:3] == "tri" and word[3] in "aeoui":
        syls += 1
    if word[:2] == "bi" and word[2] in "aeoui":
        syls += 1

    # 10)
    if word[-3:] == "ian":
        if word[-4:] == "cian" or word[-4:] == "tian":
            pass
        else:
            syls += 1

    # 11)
    if word[:2] == "co" and word[2] in 'eaoui':
        if word[:4] in co_two or word[:5] in co_two or word[:6] in co_two:
            syls += 1
        elif word[:4] in co_one or word[:5] in co_one or word[:6] in co_one:
            pass
        else:
            syls += 1

    # 12)
    if word[:3] == "pre" and word[3] in 'eaoui':
        if word[:6] in pre_one:
            pass
        else:
            syls += 1

    # 13)
    negative = ["doesn't", "isn't", "shouldn't", "couldn't", "wouldn't"]
    if word[-3:] == "n't":
        if word in negative:
            syls += 1
        else:
            pass

    # 14)
    if word in exception_del:
        disc += 1
    if word in exception_add:
        syls += 1

    # Подсчет выхода
    return (numVowels - disc + syls)

# Проходимся по каждой строке в столбце, заводим счетчик для записи\чтения в один проход
for i, row in enumerate(sheet.iter_rows(min_col=2, min_row=1, max_col=2, values_only=True), start=1): #Координаты ячеек 2-ч курсоров.
    for cell in row:
        language = detect(cell)
        if language == 'ru':
            dic = pyphen.Pyphen(lang='ru')
            words = cell.split() 
            syllables_count = 0
            for word in words: 
                syllables = dic.inserted(word).split('-')
                syllables_count += len(syllables)             
            sheet.cell(row=i, column=3, value=syllables_count)
        elif language == 'en':
            result = sylco(cell)
            sheet.cell(row=i, column=3, value=result)

# Сохраняем изменения в файле Excel
workbook.save("result.xlsx")
print("Запись завершена")